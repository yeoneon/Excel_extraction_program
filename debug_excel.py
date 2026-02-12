
import os
import pandas as pd
import openpyxl
from logger import logger

def debug_excel_structure(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb["CORSIA"] if "CORSIA" in wb.sheetnames else wb.active
    
    print(f"--- Merged Ranges for {file_path} ---")
    for range_ in ws.merged_cells.ranges:
        print(range_.coord)
        
    print(f"\n--- Style Info for Problematic Cells ---")
    cells = ['F14', 'F16', 'F17', 'F18', 'F19']
    for coord in cells:
        cell = ws[coord]
        print(f"Cell {coord}: Border={cell.border}")

if __name__ == "__main__":
    # Check the template
    template_path = r"C:/Users/LG/Desktop/기타자료/기타잡무/성기도움/Self+Declaration(UCO)_ISCC+EU_V2.3_CORSIA_V2.1_ENGLISH-KOREAN_BK에너지.xlsx"
    debug_excel_structure(template_path)
