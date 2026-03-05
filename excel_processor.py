import os
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from datetime import datetime
import random
from logger import logger

class ExcelHandler:
    def __init__(self, source_path, form_path, signature_dir, api_handler):
        self.source_path = source_path
        self.form_path = form_path
        self.signature_dir = signature_dir
        self.api_handler = api_handler
        logger.info("ExcelHandler initialized")

    def _safe_write(self, ws, coordinate, value):
        """Writes value to a cell, handling merged cells by finding the top-left coordinate."""
        from openpyxl.cell.cell import MergedCell
        cell = ws[coordinate]
        if isinstance(cell, MergedCell):
            logger.debug(f"Cell {coordinate} is a MergedCell. Finding root...")
            for range_ in ws.merged_cells.ranges:
                if coordinate in range_:
                    root_coord = range_.coord.split(':')[0]
                    logger.debug(f"Found root for {coordinate}: {root_coord}")
                    ws[root_coord].value = value
                    return
        cell.value = value

    def _safe_read(self, ws, coordinate):
        """Reads value from a cell, handling merged cells by finding the top-left coordinate."""
        from openpyxl.cell.cell import MergedCell
        cell = ws[coordinate]
        if isinstance(cell, MergedCell):
            for range_ in ws.merged_cells.ranges:
                if coordinate in range_:
                    root_coord = range_.coord.split(':')[0]
                    return ws[root_coord].value
        return cell.value

    def _format_date(self, raw_date):
        if pd.isna(raw_date):
            return "", "00000000"
        if isinstance(raw_date, (datetime, pd.Timestamp)):
            return raw_date.strftime("%Y-%m-%d"), raw_date.strftime("%Y%m%d")
        d_str = str(raw_date).replace(".", "-").replace("/", "-")
        return d_str, d_str.replace("-", "")

    def _add_signature(self, ws, row=21, col=4):
        try:
            sig_files = [f for f in os.listdir(self.signature_dir) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
            if sig_files:
                img_path = os.path.join(self.signature_dir, random.choice(sig_files))
                img = XLImage(img_path)
                img.width, img.height = 200, 33
                
                # Default for ExcelHandler is E22 (col 4, row 21)
                marker = AnchorMarker(col=col, colOff=10 * 12700, row=row, rowOff=3 * 12700)
                size = XDRPositiveSize2D(cx=img.width * 9525, cy=img.height * 9525)
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                
                ws.add_image(img)
                logger.debug(f"Added signature at row={row+1}, col={chr(65+col)}: {img_path}")
            else:
                logger.warning("No signature files found in directory.")
        except Exception as e:
            logger.error(f"Failed to add signature: {e}")

    def _reinforce_borders(self, ws):
        """Reinforces right border for specific cells that lose styling."""
        from openpyxl.styles.borders import Border, Side
        thin = Side(border_style="thin", color="000000")
        
        # Target cells: F14, F16, F17, F18, F19
        target_cells = ['F14', 'F16', 'F17', 'F18', 'F19']
        
        for coord in target_cells:
            cell = ws[coord]
            existing = cell.border
            new_border = Border(
                left=existing.left,
                right=thin,  # Enforce right border
                top=existing.top,
                bottom=existing.bottom,
                diagonal=existing.diagonal,
                diagonal_direction=existing.diagonal_direction,
                outline=existing.outline,
                vertical=existing.vertical,
                horizontal=existing.horizontal
            )
            cell.border = new_border
            logger.debug(f"Reinforced right border for {coord}")

        # Reinforce top border for B22 as requested by user
        cell_b22 = ws['B22']
        existing_b22 = cell_b22.border
        new_border_b22 = Border(
            left=existing_b22.left,
            right=existing_b22.right,
            top=thin,  # Enforce top border
            bottom=existing_b22.bottom,
            diagonal=existing_b22.diagonal,
            diagonal_direction=existing_b22.diagonal_direction,
            outline=existing_b22.outline,
            vertical=existing_b22.vertical,
            horizontal=existing_b22.horizontal
        )
        cell_b22.border = new_border_b22
        logger.debug("Reinforced top border for B22")

    def process(self, convert_pdf=True, progress_callback=None):
        """Original processing logic (NCP + Kakao style)."""
        try:
            today_str = datetime.now().strftime("%Y-%m-%d")
            output_dir = os.path.join(os.getcwd(), today_str)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                logger.info(f"Created output directory: {output_dir}")

            logger.info(f"Reading source file: {self.source_path}")
            df = pd.read_excel(self.source_path, header=5)
            logger.info(f"Total rows to process: {len(df)}")

            processed_count = 0
            total_rows = len(df)
            for index, row in df.iterrows():
                if progress_callback:
                    progress_callback(index, total_rows, f"Processing row {index + 1}/{total_rows}...")
                try:
                    logger.info(f"Processing row {index + 1}...")
                    
                    raw_date = row.iloc[1]
                    date_val, date_filename = self._format_date(raw_date)
                    if date_val == "":
                        logger.warning(f"Skipping row {index + 1}: Invalid or missing date.")
                        continue
                    representative = str(row.iloc[2]).strip()
                    company_name = str(row.iloc[3]).strip()
                    address_ko = str(row.iloc[4]).strip()
                    weight = str(row.iloc[6]).strip()

                    phone, zip_code, longitude, latitude = self.api_handler.get_enriched_data(address_ko, company_name)
                    
                    wb = openpyxl.load_workbook(self.form_path)
                    ws = wb["CORSIA"] if "CORSIA" in wb.sheetnames else wb.active

                    self._safe_write(ws, 'C4', company_name)
                    self._safe_write(ws, 'C5', address_ko)
                    self._safe_write(ws, 'C7', zip_code)
                    self._safe_write(ws, 'C8', phone)
                    self._safe_write(ws, 'B12', f"수거일 : {date_val}")
                    self._safe_write(ws, 'C12', f"수거량 : {weight} kg")
                    self._safe_write(ws, 'B13', f"DATE : {date_val}")
                    self._safe_write(ws, 'C13', f"Quantity collected: {weight} kg")
                    self._safe_write(ws, 'A22', f"{company_name}/{date_filename}")
                    self._safe_write(ws, 'C22', representative)

                    self._add_signature(ws)
                    self._reinforce_borders(ws)

                    template_name = os.path.splitext(os.path.basename(self.form_path))[0]
                    base_filename = f"{template_name}_{company_name}"
                    
                    if not phone:
                        base_filename = f"전화번호_{base_filename}"
                    if not zip_code:
                        base_filename = f"우편_{base_filename}"

                    save_filename = f"{base_filename}.xlsx"
                    save_path = os.path.join(output_dir, save_filename)

                    counter = 1
                    while os.path.exists(save_path):
                        save_filename = f"{base_filename}_{counter}.xlsx"
                        save_path = os.path.join(output_dir, save_filename)
                        counter += 1
                    
                    wb.save(save_path)
                    logger.info(f"Saved: {save_path}")
                    processed_count += 1
                    
                except Exception as row_error:
                    logger.error(f"Error in row {index + 1}: {row_error}", exc_info=True)
                    continue

            logger.info(f"Processing complete. {processed_count} files generated.")
            return processed_count, today_str

        except Exception as e:
            logger.critical(f"Critical error in ExcelHandler: {e}", exc_info=True)
            raise e

class KakaoExcelHandler(ExcelHandler):
    def __init__(self, source_path, form_path, signature_dir, api_handler):
        super().__init__(source_path, form_path, signature_dir, api_handler)
        logger.info("KakaoExcelHandler initialized")

    def _reinforce_kakao_borders(self, ws):
        """Reinforces the right border specifically for the Kakao template (I-column), handling merged cells."""
        from openpyxl.styles.borders import Border, Side
        thin = Side(border_style="thin", color="000000")
        target_cells = ['I18', 'I21', 'I22', 'I23', 'I33', 'I34']
        
        for coord in target_cells:
            # Find the actual range if merged
            target_range = None
            for r in ws.merged_cells.ranges:
                if coord in r:
                    target_range = r
                    break
            
            if target_range:
                # Apply ONLY to the right edge of the merged range
                for row in range(target_range.min_row, target_range.max_row + 1):
                    cell = ws.cell(row=row, column=target_range.max_col)
                    e = cell.border
                    cell.border = Border(
                        left=e.left,
                        right=thin,
                        top=e.top,
                        bottom=e.bottom
                    )
            else:
                # Single cell - Only enforce right border
                cell = ws[coord]
                e = cell.border
                cell.border = Border(
                    left=e.left,
                    right=thin,
                    top=e.top,
                    bottom=e.bottom
                )
            logger.debug(f"Reinforced Kakao right border for {coord}")

    def _convert_to_pdf_batch(self, excel_paths, pdf_dir=None, progress_callback=None, start_offset=0, total_steps=0):
        """Processes multiple Excel files for PDF conversion in a single instance where possible."""
        if not excel_paths:
            return
        
        import sys
        if sys.platform == "win32":
            logger.info("Starting batch PDF conversion on Windows via Excel COM...")
            excel = None
            try:
                import win32com.client
                import pythoncom
                pythoncom.CoInitialize()
                # Use DispatchEx to ensure we use isolated instance
                excel = win32com.client.DispatchEx("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False
                excel.Interactive = False
                
                for i, excel_path in enumerate(excel_paths):
                    if progress_callback:
                        progress_callback(start_offset + i, total_steps, f"Converting PDF {i+1}/{len(excel_paths)}...")
                    try:
                        abs_excel_path = os.path.abspath(excel_path)
                        if pdf_dir:
                            base_name = os.path.splitext(os.path.basename(excel_path))[0]
                            pdf_path = os.path.join(pdf_dir, f"{base_name}.pdf")
                        else:
                            pdf_path = os.path.splitext(excel_path)[0] + ".pdf"
                        
                        abs_pdf_path = os.path.abspath(pdf_path)
                        
                        if os.path.exists(abs_pdf_path):
                            try: os.remove(abs_pdf_path)
                            except: pass
                            
                        wb = excel.Workbooks.Open(abs_excel_path)
                        ws = wb.Worksheets(1)
                        ws.ExportAsFixedFormat(0, abs_pdf_path) # 0 = xlTypePDF
                        wb.Close(False)
                        logger.info(f"PDF saved: {pdf_path}")
                    except Exception as row_e:
                        logger.error(f"Failed to convert {excel_path}: {row_e}")
                
                excel.Quit()
                pythoncom.CoUninitialize()
                return True
            except Exception as e:
                logger.error(f"Critical error in batch PDF conversion: {e}")
                if excel:
                    try: excel.Quit()
                    except: pass
                return False
        else:
            # For non-Windows, fall back to individual calls or implement batching if needed
            logger.info(f"Starting PDF conversion for {len(excel_paths)} files...")
            for i, path in enumerate(excel_paths):
                if progress_callback:
                    progress_callback(start_offset + i, total_steps, f"Converting PDF {i+1}/{len(excel_paths)}...")
                self._save_as_pdf(path, pdf_dir=pdf_dir)
            return True

    def _save_as_pdf(self, excel_path, pdf_dir=None):
        """Attempts to save a single Excel file as PDF."""
        # This keeps the original logic for single-file scenarios (like NCP style if ever needed)
        if pdf_dir:
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            pdf_path = os.path.join(pdf_dir, f"{base_name}.pdf")
        else:
            base_name = os.path.splitext(excel_path)[0]
            pdf_path = f"{base_name}.pdf"
        
        import sys
        if sys.platform == "win32":
            return self._convert_to_pdf_batch([excel_path], pdf_dir=pdf_dir)
        else:
            try:
                import subprocess
                # Attempt 1: Using AppleScript (Microsoft Excel required)
                script = f'''
                set inputPath to POSIX file "{excel_path}"
                set outputPath to POSIX file "{pdf_path}"
                tell application "Microsoft Excel"
                    open inputPath
                    save sheet 1 of active workbook in outputPath as PDF file format
                    close active workbook saving no
                end tell
                '''
                result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True)
                if result.returncode == 0:
                    logger.info(f"PDF saved successfully: {pdf_path}")
                    return True
                else:
                    logger.warning(f"AppleScript failed (Excel might not be installed): {result.stderr}")
                    
                # Attempt 2: Using LibreOffice if available
                res = subprocess.run(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', os.path.dirname(excel_path), excel_path], capture_output=True)
                if res.returncode == 0:
                    logger.info(f"PDF saved via LibreOffice: {pdf_path}")
                    return True
                    
            except Exception as e:
                logger.error(f"PDF conversion error: {e}")
        
        logger.warning("PDF conversion failed. Falling back to Excel only.")
        return False

    def process(self, convert_pdf=True, progress_callback=None):
        """Kakao-only specific processing logic."""
        try:
            today_str = datetime.now().strftime("%Y-%m-%d")
            base_dir = os.path.join(os.getcwd(), today_str)
            success_excel_dir = os.path.join(base_dir, "성공엑셀")
            no_phone_dir = os.path.join(base_dir, "전화번호미기입 엑셀")
            address_error_dir = os.path.join(base_dir, "주소오류 엑셀")
            pdf_dir = os.path.join(base_dir, "완성PDF")
            
            for d in [success_excel_dir, no_phone_dir, address_error_dir, pdf_dir]:
                if not os.path.exists(d):
                    os.makedirs(d)

            logger.info(f"Reading source file: {self.source_path}")
            df = pd.read_excel(self.source_path, header=5)
            logger.info(f"Total rows: {len(df)}")

            processed_count = 0
            excel_conversion_list = []
            total_rows = len(df)
            
            for index, row in df.iterrows():
                if progress_callback:
                    # Initial steps: rows / total (Excel + PDF)
                    total_steps = total_rows + (total_rows if convert_pdf else 0)
                    progress_callback(index, total_steps, f"Generating Excel {index + 1}/{total_rows}...")
                try:
                    logger.info(f"Processing row {index + 1}...")
                    
                    raw_date = row.iloc[1]
                    date_val, date_filename = self._format_date(raw_date)
                    if date_val == "":
                        logger.warning(f"Skipping row {index + 1}: Invalid or missing date.")
                        continue
                    representative = str(row.iloc[2]).strip()
                    company_name = str(row.iloc[3]).strip()
                    address_ko = str(row.iloc[4]).strip()
                    weight = str(row.iloc[6]).strip()

                    # Kakao Enrichment
                    phone, zip_code, longitude, latitude = self.api_handler.get_kakao_data(address_ko, company_name)
                    
                    wb = openpyxl.load_workbook(self.form_path)
                    ws = wb.active # Use the first sheet

                    # C4: 상호명, C5: 주소(국문), C7: 대한민국 Republic of Korea, C8: 경도, F8: 위도, C9: 전화번호
                    self._safe_write(ws, 'C4', company_name)
                    self._safe_write(ws, 'C5', address_ko)
                    self._safe_write(ws, 'C7', "대한민국 Republic of Korea")
                    self._safe_write(ws, 'C8', latitude)
                    self._safe_write(ws, 'F8', longitude)
                    self._safe_write(ws, 'C9', phone)
                    
                    # A15: Dynamic template-based replacement for {대표자}
                    template_text = self._safe_read(ws, 'A15')
                    if isinstance(template_text, str) and "{대표자}" in template_text:
                        bold_large_font = InlineFont(b=True, sz=14)
                        normal_large_font = InlineFont(sz=12)
                        parts = template_text.split("{대표자}")
                        rich_text_elements = []
                        for i, part in enumerate(parts):
                            if part:
                                rich_text_elements.append(TextBlock(normal_large_font, part))
                            if i < len(parts) - 1:
                                rich_text_elements.append(TextBlock(bold_large_font, representative))
                        
                        rich_text = CellRichText(rich_text_elements)
                        self._safe_write(ws, 'A15', rich_text)
                    else:
                        # Fallback to the hardcoded bilingual text if template doesn't have the tag
                        bold_large_font = InlineFont(b=True, sz=14)
                        normal_large_font = InlineFont(sz=12)
                        rich_text = CellRichText([
                            TextBlock(normal_large_font, "By signing this self-declaration, I, "),
                            TextBlock(bold_large_font, representative),
                            TextBlock(normal_large_font, ", acting in my capacity as 담당자 and authorised representative of the Point of Origin, hereby declare, confirm and agree to the following on behalf of the Point of Origin:\n"),
                            TextBlock(normal_large_font, "본 자가선언서에 서명함으로써, 본인 "),
                            TextBlock(bold_large_font, representative),
                            TextBlock(normal_large_font, "는 담당자 의 직책으로서 Point of Origin의 권한 있는 대표로서 다음 사항을 Point of Origin 대신하여 선언하고, 확인하며, 이에 동의합니다.")
                        ])
                        self._safe_write(ws, 'A15', rich_text)
                    
                    # A16: 수거일, A17: DATE, C16: 수거량, C17: Quantity collected
                    self._safe_write(ws, 'A16', f"수거일 : {date_val}")
                    self._safe_write(ws, 'A17', f"DATE : {date_val}")
                    self._safe_write(ws, 'C16', f"수거량 : {weight}kg")
                    self._safe_write(ws, 'C17', f"Quantity collected : {weight}kg")
                    
                    # A35: 상호명/yyyyMMdd, D35: 대표자명/배출담당자, G35: 사인
                    self._safe_write(ws, 'A35', f"{company_name} / {date_filename}")
                    self._safe_write(ws, 'D35', f"{representative}/배출담당자")
                    
                    # Signature at G35 (col 6, row 34)
                    self._add_signature(ws, row=34, col=6)
                    
                    # Reinforce template-specific borders
                    self._reinforce_kakao_borders(ws)

                    # Save Logic
                    base_name = f"{company_name}_{date_filename}"
                    
                    # Determine target directory
                    is_success = bool(phone and zip_code and longitude and latitude)
                    if is_success:
                        target_dir = success_excel_dir
                    elif not phone:
                        target_dir = no_phone_dir
                    else:
                        target_dir = address_error_dir
                    
                    save_filename = f"{base_name}.xlsx"
                    save_path = os.path.join(target_dir, save_filename)
                    
                    counter = 1
                    while os.path.exists(save_path):
                        save_filename = f"{base_name}_{counter}.xlsx"
                        save_path = os.path.join(target_dir, save_filename)
                        counter += 1
                        
                    wb.save(save_path)
                    if is_success:
                        logger.info(f"Saved Excel: {save_path}")
                        # Defer PDF conversion only for successful cases
                        excel_conversion_list.append(save_path)
                    elif not phone:
                        logger.info(f"Saved Excel (No Phone): {save_path}")
                    else:
                        logger.info(f"Saved Excel (Address Error): {save_path}")
                    
                    processed_count += 1
                    
                except Exception as row_error:
                    logger.error(f"Error in row {index + 1}: {row_error}", exc_info=True)
                    continue

            # Perform PDF conversion after all Excel files are generated
            if convert_pdf and excel_conversion_list:
                total_steps = total_rows + len(excel_conversion_list)
                self._convert_to_pdf_batch(excel_conversion_list, pdf_dir=pdf_dir, progress_callback=progress_callback, start_offset=total_rows, total_steps=total_steps)

            if progress_callback:
                progress_callback(total_steps, total_steps, "Complete!")

            logger.info(f"Processing complete. {processed_count} files generated.")
            return processed_count, today_str

        except Exception as e:
            logger.critical(f"Critical error in KakaoExcelHandler: {e}", exc_info=True)
            raise e
